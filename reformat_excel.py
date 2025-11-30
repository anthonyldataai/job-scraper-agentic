"""
Script to reformat existing Excel output with proper hyperlinks and column widths
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def reformat_excel(input_file, output_file=None):
    if output_file is None:
        output_file = input_file.replace('.xlsx', '_formatted.xlsx')
    
    # Read the Excel file
    df = pd.read_excel(input_file)
    
    # Save to new file
    df.to_excel(output_file, index=False)
    
    # Apply formatting
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Set column widths and formatting
    for column_cells in ws.columns:
        col_letter = column_cells[0].column_letter
        header_value = ws.cell(row=1, column=column_cells[0].column).value
        
        if header_value == "Link":
            ws.column_dimensions[col_letter].width = 80
        elif header_value == "match_reasoning":
            ws.column_dimensions[col_letter].width = 100
            # Apply text wrap to reasoning column
            for cell in column_cells[1:]:  # Skip header
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        else:
            # Auto-fit other columns
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[col_letter].width = min(length + 2, 50)
    
    # Make Link column clickable with blue underlined style
    link_col_idx = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Link":
            link_col_idx = col
            break
    
    if link_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col_idx)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0000FF", underline="single")
    
    wb.save(output_file)
    print(f"Reformatted Excel saved to: {output_file}")
    return output_file

if __name__ == "__main__":
    # Reformat the latest output file
    input_file = "new_leads_2025-11-28_14-53.xlsx"
    output_file = reformat_excel(input_file)
