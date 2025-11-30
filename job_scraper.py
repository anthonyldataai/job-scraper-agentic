import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import urllib.parse
import os
from datetime import datetime, timedelta
import re
from utils.persistence import log_agent_action, get_config_value
import json
import random

def parse_relative_date(date_str):
    """Parses relative date strings like '3 days ago' or '17 October' into a datetime object."""
    if not date_str or date_str == "N/A":
        return None
    
    today = datetime.now()
    date_str = date_str.lower().strip()
    
    # Clean up common prefixes like "published:", "posted:", "active:"
    date_str = re.sub(r'^(published|posted|active):\s*', '', date_str)
    date_str = date_str.strip()
    
    try:
        if "just posted" in date_str or "today" in date_str or date_str == "0 days ago":
            return today
        
        if "yesterday" in date_str:
            return today - timedelta(days=1)
        
        # Regex for "X days/weeks/months ago"
        match = re.search(r'(\d+)\s+(day|week|month|hour|minute)s?\s+ago', date_str)
        if match:
            num = int(match.group(1))
            unit = match.group(2)
            
            if unit == "day":
                return today - timedelta(days=num)
            elif unit == "week":
                return today - timedelta(weeks=num)
            elif unit == "month":
                return today - timedelta(days=num*30)
            elif unit == "hour":
                return today - timedelta(hours=num)
            elif unit == "minute":
                return today - timedelta(minutes=num)
        
        # Handle dates like "17 October" or "21 November"
        match = re.search(r'(\d+)\s+(january|february|march|april|may|june|july|august|september|october|november|december)', date_str)
        if match:
            day = int(match.group(1))
            month_name = match.group(2)
            
            # Map month names to numbers
            months = {
                'january': 1, 'february': 2, 'march': 3, 'april': 4,
                'may': 5, 'june': 6, 'july': 7, 'august': 8,
                'september': 9, 'october': 10, 'november': 11, 'december': 12
            }
            month = months[month_name]
            
            # Assume current year, but if the date is in the future, use previous year
            year = today.year
            parsed_date = datetime(year, month, day)
            if parsed_date > today:
                parsed_date = datetime(year - 1, month, day)
            
            return parsed_date
        
        return None
    except:
        return None

async def scrape_indeed(page, query, limit=None):
    print(f"Scraping Indeed for: {query}")
    encoded_query = urllib.parse.quote(query)
    base_url = f"https://uk.indeed.com/jobs?q={encoded_query}&l=London&sort=date"
    
    all_jobs = []
    
    # If limit is small, only scrape page 1
    pages_to_scrape = [0, 10, 20]
    if limit and limit <= 10:
        pages_to_scrape = [0]
    
    for start in pages_to_scrape:
        print(f"  Indeed Page {start//10 + 1}...")
        url = f"{base_url}&start={start}"
        
        try:
            await page.goto(url, timeout=60000)
            try:
                await page.wait_for_selector('.job_seen_beacon', timeout=10000)
            except:
                print("  Indeed: No jobs found or captcha on this page.")
                break
            
            job_cards = await page.query_selector_all('.job_seen_beacon')
            if not job_cards:
                break

            for card in job_cards:
                if limit and len(all_jobs) >= limit:
                    break
                    
                try:
                    title_el = await card.query_selector('h2.jobTitle span')
                    company_el = await card.query_selector('[data-testid="company-name"]')
                    location_el = await card.query_selector('[data-testid="text-location"]')
                    link_el = await card.query_selector('h2.jobTitle a')
                    date_el = await card.query_selector('[data-testid="myJobsStateDate"]')
                    if not date_el:
                        date_el = await card.query_selector('.date')
                    salary_el = await card.query_selector('.salary-snippet-container')
                    
                    title = await title_el.inner_text() if title_el else "N/A"
                    company = await company_el.inner_text() if company_el else "N/A"
                    location = await location_el.inner_text() if location_el else "N/A"
                    link_suffix = await link_el.get_attribute('href') if link_el else ""
                    link = f"https://uk.indeed.com{link_suffix}" if link_suffix else "N/A"
                    posted_date_str = await date_el.inner_text() if date_el else "N/A"
                    posted_date_str = posted_date_str.replace("Active ", "").strip()
                    salary = await salary_el.inner_text() if salary_el else "N/A"
                    
                    all_jobs.append({
                        "Title": title,
                        "Company": company,
                        "Location": location,
                        "Link": link,
                        "Posted Date Text": posted_date_str,
                        "Posted Date": parse_relative_date(posted_date_str),
                        "Salary": salary,
                        "Applicants": "N/A",
                        "Job Type": "N/A",
                        "Apply Method": "Apply",
                        "Source": "Indeed"
                    })
                except Exception as e:
                    continue
            
            if limit and len(all_jobs) >= limit:
                break
                
            await asyncio.sleep(2)
        except Exception as e:
            print(f"Error scraping Indeed page: {e}")
            break
            
    print(f"Found {len(all_jobs)} jobs on Indeed.")
    return all_jobs

async def scrape_totaljobs_details(page, link):
    """Visits a TotalJobs job page and uses LLM to extract details intelligently."""
    try:
        await page.goto(link, timeout=30000)
        try:
            await page.wait_for_selector('h1', timeout=5000)
        except:
            pass

        # Get the page content
        page_content = await page.content()
        
        # Use LLM to extract job details
        from utils.llm_extractor import extract_job_details_with_llm
        extracted_data = await extract_job_details_with_llm(page_content, link)
        
        return (
            extracted_data.get("company", "N/A"),
            extracted_data.get("salary", "N/A"),
            extracted_data.get("location", "N/A"),
            extracted_data.get("job_type", "N/A"),
            extracted_data.get("posted_date", "N/A")
        )
        
    except Exception as e:
        print(f"Error extracting job details with LLM: {e}")
        return "N/A", "N/A", "N/A", "N/A", "N/A"

async def scrape_totaljobs(page, query):
    print(f"Scraping TotalJobs for: {query}")
    encoded_query = urllib.parse.quote(query)
async def scrape_totaljobs(page, query, limit=None):
    print(f"Scraping TotalJobs for: {query}")
    encoded_query = urllib.parse.quote(query)
    base_url = f"https://www.totaljobs.com/jobs/{encoded_query}/in-london?radius=10&postedwithin=7"
    
    all_jobs = []
    
    # If limit is small, only scrape page 1
    pages_to_scrape = range(1, 4)
    if limit and limit <= 25:
        pages_to_scrape = [1]
        
    for page_num in pages_to_scrape:
        print(f"  TotalJobs Page {page_num}...")
        url = base_url if page_num == 1 else f"{base_url}/page-{page_num}"
    
        try:
            await page.goto(url, timeout=60000)
            try:
                await page.wait_for_selector('a[href*="/job/"]', timeout=10000)
            except:
                print("  TotalJobs: No jobs found on this page.")
                break
            
            job_links = await page.query_selector_all('a[href*="/job/"]')
            if not job_links:
                break
            
            page_jobs = []
            for link_el in job_links:
                if limit and len(all_jobs) + len(page_jobs) >= limit:
                    break
                    
                try:
                    title = await link_el.inner_text()
                    href = await link_el.get_attribute('href')
                    
                    if not title or not href:
                        continue

                    link = f"https://www.totaljobs.com{href}" if href.startswith("/") else href

                    if any(j['Link'] == link for j in all_jobs):
                        continue

                    page_jobs.append({
                        "Title": title.strip(),
                        "Company": "N/A",  # Will be filled from detail page
                        "Location": "N/A",  # Will be filled from detail page
                        "Link": link,
                        "Posted Date Text": "N/A",  # Will be filled from detail page
                        "Posted Date": None,
                        "Salary": "N/A",
                        "Applicants": "N/A",
                        "Job Type": "N/A",  # Will be filled from detail page
                        "Apply Method": "Apply",
                        "Source": "TotalJobs"
                    })
                except Exception as e:
                    continue
            
            if not page_jobs:
                break
            
            print(f"  Found {len(page_jobs)} jobs on this page. Fetching details...")
            
            # Visit each job page to get details
            for job in page_jobs:
                if job["Link"] != "N/A":
                    company, salary, location, job_type, posted_date = await scrape_totaljobs_details(page, job["Link"])
                    job["Company"] = company
                    job["Salary"] = salary
                    job["Location"] = location
                    job["Job Type"] = job_type
                    job["Posted Date Text"] = posted_date
                    job["Posted Date"] = parse_relative_date(posted_date)
                    await asyncio.sleep(1)  # Be polite
            
            all_jobs.extend(page_jobs)
            await asyncio.sleep(2)
        except Exception as e:
            print(f"Error scraping TotalJobs page: {e}")
            break
            
    print(f"Found {len(all_jobs)} jobs on TotalJobs.")
    return all_jobs

async def scrape_cwjobs(page, query):
    print(f"Scraping CWJobs for: {query}")
    encoded_query = urllib.parse.quote(query)
    base_url = f"https://www.cwjobs.co.uk/jobs/{encoded_query}/in-london"
    
    all_jobs = []
    for page_num in range(1, 4):
        print(f"  CWJobs Page {page_num}...")
        url = base_url if page_num == 1 else f"{base_url}/page-{page_num}"
    
        try:
            await page.goto(url, timeout=60000)
            try:
                await page.wait_for_selector('a[href*="/job/"]', timeout=10000)
            except:
                print("  CWJobs: No jobs found on this page.")
                break
            
            job_links = await page.query_selector_all('a[href*="/job/"]')
            if not job_links:
                break
            
            page_jobs_count = 0
            for link_el in job_links:
                try:
                    title = await link_el.inner_text()
                    href = await link_el.get_attribute('href')
                    
                    if not title or not href:
                        continue

                    link = f"https://www.cwjobs.co.uk{href}" if href.startswith("/") else href
                    card = await link_el.evaluate_handle('el => el.closest("div[class*=\'res-\']")')
                    card_text = await card.inner_text() if card else ""
                    
                    posted_date_str = "N/A"
                    date_match = re.search(r'(Posted|Active)\s+(\d+\s+\w+s?\s+ago|today|yesterday)', card_text, re.IGNORECASE)
                    if date_match:
                        posted_date_str = date_match.group(2)
                    elif "Today" in card_text:
                        posted_date_str = "Today"
                    elif "Yesterday" in card_text:
                        posted_date_str = "Yesterday"

                    salary = "N/A"
                    salary_match = re.search(r'£[\d,]+(\s*-\s*£[\d,]+)?', card_text)
                    if salary_match:
                        salary = salary_match.group(0)

                    job_type = "N/A"
                    if "Remote" in card_text:
                        job_type = "Remote"
                    elif "Hybrid" in card_text:
                        job_type = "Hybrid"
                    elif "Permanent" in card_text:
                        job_type = "Permanent"

                    if any(j['Link'] == link for j in all_jobs):
                        continue

                    all_jobs.append({
                        "Title": title.strip(),
                        "Company": "N/A",
                        "Location": "N/A",
                        "Link": link,
                        "Posted Date Text": posted_date_str,
                        "Posted Date": parse_relative_date(posted_date_str),
                        "Salary": salary,
                        "Applicants": "N/A",
                        "Job Type": job_type,
                        "Apply Method": "Apply",
                        "Source": "CWJobs"
                    })
                    page_jobs_count += 1
                except Exception as e:
                    continue
            
            if page_jobs_count == 0:
                break
                
            await asyncio.sleep(2)
        except Exception as e:
            print(f"Error scraping CWJobs page: {e}")
            break
            
    print(f"Found {len(all_jobs)} jobs on CWJobs.")
    return all_jobs

async def scrape_reed_details(page, link):
    """Visits a Reed job page to extract details."""
    try:
        await page.goto(link, timeout=30000)
        try:
            await page.wait_for_selector('h1', timeout=5000)
        except:
            pass

        # Extract posted date and company (e.g., "Yesterday by Grant Thornton" or "17 October by Gold Group Ltd")
        posted_date_str = "N/A"
        company = "N/A"
        try:
            # Look for the date/company line - avoid salary patterns with £
            page_text = await page.content()
            
            # Try to find elements with date information
            date_elements = await page.query_selector_all('span, div, time')
            for el in date_elements:
                text = await el.inner_text()
                text = text.strip()
                
                # Skip if it contains £ (salary)
                if '£' in text:
                    continue
                
                # Look for "Yesterday by Company" or "17 October by Company"
                match = re.search(r'(yesterday|\d+\s+\w+)\s+by\s+(.+)', text, re.IGNORECASE)
                if match:
                    posted_date_str = match.group(1).strip()
                    company = match.group(2).strip()
                    break
                
                # Look for just date patterns (without "by")
                if not match and len(text) < 50:  # Short text only
                    # Check for "Yesterday", "17 October", etc.
                    if re.search(r'(yesterday|today|\d+\s+(january|february|march|april|may|june|july|august|september|october|november|december))', text, re.IGNORECASE):
                        if 'by' not in text.lower():
                            posted_date_str = text
        except:
            pass

        # Extract job type (e.g., "Contract, full-time")
        job_type = "N/A"
        try:
            # Look for job type keywords
            page_text = await page.content()
            if re.search(r'Permanent,?\s+full-time', page_text, re.IGNORECASE):
                job_type = "Permanent, full-time"
            elif 'Permanent' in page_text:
                job_type = "Permanent"
            elif re.search(r'Contract,?\s+full-time', page_text, re.IGNORECASE):
                job_type = "Contract, full-time"
            elif 'Contract' in page_text:
                job_type = "Contract"
        except:
            pass
        
        return posted_date_str, company, job_type
    except Exception as e:
        return "N/A", "N/A", "N/A"

async def scrape_reed(page, query, limit=None):
    print(f"Scraping Reed for: {query}")
    encoded_keywords = urllib.parse.quote(query)
    base_url = f"https://www.reed.co.uk/jobs?keywords={encoded_keywords}&location=London&sortby=DisplayDate"
    
    all_jobs = []
    
    # If limit is small, only scrape page 1
    pages_to_scrape = range(1, 4)
    if limit and limit <= 25:
        pages_to_scrape = [1]
        
    for page_num in pages_to_scrape:
        print(f"  Reed Page {page_num}...")
        url = base_url if page_num == 1 else f"{base_url}&pageno={page_num}"
    
        try:
            await page.goto(url, timeout=60000)
            
            if page_num == 1:
                try:
                    await asyncio.sleep(2)
                    reject_btn = await page.query_selector('button:has-text("Reject All")')
                    if reject_btn:
                        print("  Clicking 'Reject All' on Reed...")
                        await reject_btn.click()
                        await asyncio.sleep(2)
                except Exception as e:
                    print(f"  Cookie dialog handling: {e}")
            
            await asyncio.sleep(2)
            
            try:
                await page.wait_for_selector('article.job-result, div[class*="job-card"], div[data-qa="job-card"]', timeout=10000)
            except:
                print("  Reed: No jobs found on this page.")
                if page_num == 1:
                    await page.screenshot(path="reed_page1_debug.png")
                    print("  Screenshot saved to reed_page1_debug.png")
                break
            
            job_cards = await page.query_selector_all('article.job-result')
            if not job_cards:
                job_cards = await page.query_selector_all('div[class*="job-card"]')
            if not job_cards:
                job_cards = await page.query_selector_all('div[data-qa="job-card"]')
            
            if not job_cards:
                print(f"  No job cards found with any selector")
                break
            
            print(f"  Found {len(job_cards)} job cards")
            
            page_jobs = []
            for card in job_cards:
                if limit and len(all_jobs) + len(page_jobs) >= limit:
                    break
                try:
                    title_el = await card.query_selector('h3.title a, h2 a, a[data-qa="job-card-title"]')
                    location_el = await card.query_selector('.location, span[class*="location"]')
                    salary_el = await card.query_selector('.salary, span[class*="salary"]')
                    
                    if not title_el:
                        continue
                    
                    title = await title_el.get_attribute('title')
                    if not title:
                        title = await title_el.inner_text()
                    
                    link_suffix = await title_el.get_attribute('href') if title_el else ""
                    link = f"https://www.reed.co.uk{link_suffix}" if link_suffix and link_suffix.startswith("/") else link_suffix
                    
                    if link == "N/A" or not link:
                        continue
                    
                    location = await location_el.inner_text() if location_el else "N/A"
                    salary = await salary_el.inner_text() if salary_el else "N/A"

                    if any(j['Link'] == link for j in all_jobs):
                        continue

                    page_jobs.append({
                        "Title": title.strip(),
                        "Company": "N/A",  # Will be filled from detail page
                        "Location": location.strip(),
                        "Link": link,
                        "Posted Date Text": "N/A",  # Will be filled from detail page
                        "Posted Date": None,
                        "Salary": salary.strip(),
                        "Applicants": "N/A",
                        "Job Type": "N/A",  # Will be filled from detail page
                        "Apply Method": "Apply",
                        "Source": "Reed"
                    })
                except Exception as e:
                    continue
            
            if not page_jobs:
                break
            
            print(f"  Found {len(page_jobs)} jobs on this page. Fetching details...")
            
            # Visit each job page to get details
            for job in page_jobs:
                if job["Link"] != "N/A":
                    posted_date, company, job_type = await scrape_reed_details(page, job["Link"])
                    job["Posted Date Text"] = posted_date
                    job["Posted Date"] = parse_relative_date(posted_date)
                    job["Company"] = company
                    job["Job Type"] = job_type
                    await asyncio.sleep(1)  # Be polite
            
            all_jobs.extend(page_jobs)
            await asyncio.sleep(2)
        except Exception as e:
            print(f"Error scraping Reed page: {e}")
            break
            
    print(f"Found {len(all_jobs)} jobs on Reed.")
    return all_jobs

async def scrape_glassdoor(page, query, limit=None):
    print(f"Scraping Glassdoor for: {query}")
    encoded_query = urllib.parse.quote(query)
    base_url = f"https://www.glassdoor.co.uk/Job/jobs.htm?sc.keyword={encoded_query}&locT=C&locId=2671300&fromAge=7"
    
    all_jobs = []
    print(f"  Glassdoor Page 1...")
    
    try:
        await page.goto(base_url, timeout=60000)
        
        try:
            accept_btn = await page.query_selector('button:has-text("Accept")')
            if accept_btn:
                await accept_btn.click()
                await asyncio.sleep(1)
        except:
            pass
        
        try:
            content = await page.content()
            if "Cloudflare" in content or "Verify you are human" in content:
                print("  Glassdoor: Blocked by Cloudflare.")
                return []
                
            await page.wait_for_selector('li[data-test="jobListing"]', timeout=10000)
        except:
            print("  Glassdoor: No jobs found or blocked.")
            return []
        
        job_cards = await page.query_selector_all('li[data-test="jobListing"]')
        
        for card in job_cards:
            if limit and len(all_jobs) >= limit:
                break
            try:
                title_el = await card.query_selector('a[data-test="job-link"]')
                company_el = await card.query_selector('span.EmployerProfile_employerName__8w0oV')
                if not company_el:
                    company_el = await card.query_selector('div.EmployerProfile_employerName__8w0oV')
                
                location_el = await card.query_selector('span[data-test="emp-location"]')
                salary_el = await card.query_selector('span[data-test="detailSalary"]')
                date_el = await card.query_selector('div[data-test="job-age"]')
                
                title = await title_el.inner_text() if title_el else "N/A"
                link = await title_el.get_attribute('href') if title_el else "N/A"
                if link != "N/A" and not link.startswith("http"):
                    link = f"https://www.glassdoor.co.uk{link}"
                
                company = await company_el.inner_text() if company_el else "N/A"
                location = await location_el.inner_text() if location_el else "N/A"
                salary = await salary_el.inner_text() if salary_el else "N/A"
                posted_date_str = await date_el.inner_text() if date_el else "N/A"

                if any(j['Link'] == link for j in all_jobs):
                    continue

                all_jobs.append({
                    "Title": title.strip(),
                    "Company": company.strip(),
                    "Location": location.strip(),
                    "Link": link,
                    "Posted Date Text": posted_date_str,
                    "Posted Date": parse_relative_date(posted_date_str),
                    "Salary": salary.strip(),
                    "Applicants": "N/A",
                    "Job Type": "N/A",
                    "Apply Method": "Apply",
                    "Source": "Glassdoor"
                })
            except Exception as e:
                continue
    except Exception as e:
        print(f"Error scraping Glassdoor: {e}")
        
    print(f"Found {len(all_jobs)} jobs on Glassdoor.")
    return all_jobs

async def scrape_linkedin_details(page, link):
    """Visits a LinkedIn job page to extract details."""
    try:
        await page.goto(link, timeout=30000)
        try:
            await page.wait_for_selector('.top-card-layout__entity-info', timeout=5000)
        except:
            pass

        applicants = "N/A"
        try:
            app_el = await page.query_selector('.num-applicants__caption')
            if app_el:
                applicants = await app_el.inner_text()
            else:
                content = await page.content()
                match = re.search(r'(\d+)\s+applicants?', content)
                if match:
                    applicants = f"{match.group(1)} applicants"
        except:
            pass

        job_type = "N/A"
        try:
            badges = await page.query_selector_all('.job-details-jobs-unified-top-card__job-insight')
            for badge in badges:
                text = await badge.inner_text()
                if "Remote" in text:
                    job_type = "Remote"
                    break
                elif "Hybrid" in text:
                    job_type = "Hybrid"
                    break
                elif "On-site" in text:
                    job_type = "On-site"
                    break
        except:
            pass

        apply_method = "Apply"
        try:
            easy_apply_btn = await page.query_selector('button.jobs-apply-button--top-card')
            if easy_apply_btn:
                btn_text = await easy_apply_btn.inner_text()
                if "Easy Apply" in btn_text:
                    apply_method = "Easy Apply"
        except:
            pass
        
        return applicants.strip(), job_type, apply_method
    except Exception as e:
        return "N/A", "N/A", "Apply"

async def scrape_linkedin(page, query, limit=None):
    print(f"Scraping LinkedIn (Public) for: {query}")
    encoded_query = urllib.parse.quote(query)
    base_url = f"https://www.linkedin.com/jobs/search?keywords={encoded_query}&location=London&f_TPR=r604800"
    
    all_jobs = []
    # If limit is small (e.g. 5), we only need the first page (start=0)
    pages_to_scrape = [0, 25, 50]
    if limit and limit <= 25:
        pages_to_scrape = [0]
        
    for start in pages_to_scrape:
        print(f"  LinkedIn Page {start//25 + 1}...")
        url = f"{base_url}&start={start}"
        
        try:
            await page.goto(url, timeout=60000)
            try:
                await page.wait_for_selector('.jobs-search__results-list', timeout=10000)
            except:
                print("  LinkedIn: No jobs found or auth wall.")
                break
            
            job_cards = await page.query_selector_all('ul.jobs-search__results-list li')
            if not job_cards:
                break
            
            page_jobs = []
            for card in job_cards:
                if limit and len(all_jobs) + len(page_jobs) >= limit:
                    break
                    
                try:
                    title_el = await card.query_selector('h3.base-search-card__title')
                    company_el = await card.query_selector('h4.base-search-card__subtitle')
                    location_el = await card.query_selector('span.job-search-card__location')
                    link_el = await card.query_selector('a.base-card__full-link')
                    date_el = await card.query_selector('time.job-search-card__listdate')
                    if not date_el:
                        date_el = await card.query_selector('time.job-search-card__listdate--new')
                    
                    title = await title_el.inner_text() if title_el else "N/A"
                    company = await company_el.inner_text() if company_el else "N/A"
                    location = await location_el.inner_text() if location_el else "N/A"
                    link = await link_el.get_attribute('href') if link_el else "N/A"
                    posted_date_str = await date_el.inner_text() if date_el else "N/A"
                    
                    if "?" in link:
                        link = link.split("?")[0]

                    job_data = {
                        "Title": title.strip(),
                        "Company": company.strip(),
                        "Location": location.strip(),
                        "Link": link,
                        "Posted Date Text": posted_date_str.strip(),
                        "Posted Date": parse_relative_date(posted_date_str.strip()),
                        "Salary": "N/A",
                        "Applicants": "N/A",
                        "Job Type": "N/A",
                        "Apply Method": "Apply",
                        "Source": "LinkedIn"
                    }
                    
                    page_jobs.append(job_data)
                except Exception as e:
                    continue
            
            if not page_jobs:
                break
                
            print(f"  Found {len(page_jobs)} jobs on this page. Fetching details...")
            
            for job in page_jobs:
                if job["Link"] != "N/A":
                    applicants, job_type, apply_method = await scrape_linkedin_details(page, job["Link"])
                    job["Applicants"] = applicants
                    job["Job Type"] = job_type
                    job["Apply Method"] = apply_method
                    await asyncio.sleep(1)
            
            all_jobs.extend(page_jobs)
            if limit and len(all_jobs) >= limit:
                break
                
            await asyncio.sleep(2)
        except Exception as e:
            print(f"Error scraping LinkedIn page: {e}")
            break
            
    print(f"Found {len(all_jobs)} jobs on LinkedIn.")
    return all_jobs

async def scrape_all_jobs(test_mode=False, enabled_sources=None):
    """
    Main function to scrape all configured job boards using keywords from the database.
    enabled_sources: list of source names to scrape (e.g., ['linkedin', 'reed'])
    """
    # Fetch keywords from DB
    keywords_json = get_config_value("keywords")
    if not keywords_json:
        log_agent_action("Scraper", "No keywords found in configuration. Using default.", status="WARNING")
        keywords = ["Technical Project Manager"]
    else:
        try:
            keywords = json.loads(keywords_json)
        except json.JSONDecodeError:
            keywords = [keywords_json] # Fallback if stored as plain string

    # Fetch enabled sources from DB if not provided
    if enabled_sources is None:
        sources_json = get_config_value("enabled_sources")
        if sources_json:
            try:
                enabled_sources = json.loads(sources_json)
            except json.JSONDecodeError:
                enabled_sources = ['totaljobs', 'reed', 'glassdoor', 'linkedin']
        else:
            # Default to all sources
            enabled_sources = ['indeed', 'totaljobs', 'cwjobs', 'reed', 'glassdoor', 'linkedin']

    log_agent_action("Scraper", f"Starting scrape for keywords: {keywords} (Test Mode: {test_mode})", status="INFO")
    log_agent_action("Scraper", f"Enabled sources: {enabled_sources}", status="INFO")
    
    all_jobs = []
    limit = 5 if test_mode else None
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36"
        )
        page = await context.new_page()
        
        for keyword in keywords:
            log_agent_action("Scraper", f"Scraping for keyword: {keyword}", status="INFO")
            
            # Indeed
            if 'indeed' in enabled_sources:
                indeed_jobs = await scrape_indeed(page, keyword, limit=limit)
                all_jobs.extend(indeed_jobs)
            
            # TotalJobs
            if 'totaljobs' in enabled_sources:
                totaljobs_jobs = await scrape_totaljobs(page, keyword, limit=limit)
                all_jobs.extend(totaljobs_jobs)
            
            # CWJobs
            if 'cwjobs' in enabled_sources:
                cwjobs_jobs = await scrape_cwjobs(page, keyword)
                all_jobs.extend(cwjobs_jobs)
            
            # Reed
            if 'reed' in enabled_sources:
                reed_jobs = await scrape_reed(page, keyword, limit=limit)
                all_jobs.extend(reed_jobs)
            
            # Glassdoor
            if 'glassdoor' in enabled_sources:
                glassdoor_jobs = await scrape_glassdoor(page, keyword, limit=limit)
                all_jobs.extend(glassdoor_jobs)

            # LinkedIn
            if 'linkedin' in enabled_sources:
                linkedin_jobs = await scrape_linkedin(page, keyword, limit=limit)
                all_jobs.extend(linkedin_jobs)
            
            # Random delay between keywords to be polite
            await asyncio.sleep(random.uniform(2, 5))

        await browser.close()
        
    return all_jobs

def save_jobs_to_excel(jobs, filename):
    if not jobs:
        print("No jobs to save.")
        return

    df = pd.DataFrame(jobs)
    
    seven_days_ago = datetime.now() - timedelta(days=7)
    
    def is_recent_or_unknown(row):
        date = row['Posted Date']
        if pd.isna(date) or date == "N/A" or date is None:
            return True
        return date >= seven_days_ago

    df = df[df.apply(is_recent_or_unknown, axis=1)]
    df = df.sort_values(by='Posted Date', ascending=False, na_position='last')
    df = df.drop(columns=['Posted Date'])
    df = df.rename(columns={'Posted Date Text': 'Posted Date'})
    
    print(f"Total jobs to save (Recent + Unknown): {len(df)}")
    
    df.to_excel(filename, index=False)
    
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb = load_workbook(filename)
    ws = wb.active
    
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        col_letter = column_cells[0].column_letter
        
        if ws.cell(row=1, column=column_cells[0].column).value == "Link":
            ws.column_dimensions[col_letter].width = 80
        else:
            ws.column_dimensions[col_letter].width = min(length + 2, 50)

    link_col_idx = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Link":
            link_col_idx = col
            break
    
    if link_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col_idx)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0000FF", underline="single")
    
    wb.save(filename)
    print(f"Saved to {filename}")

if __name__ == "__main__":
    jobs = asyncio.run(scrape_all_jobs())
    save_jobs_to_excel(jobs, "jobs_v4.xlsx")
